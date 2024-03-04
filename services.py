import os
import shutil
import threading
import time
from functools import cache
from urllib.parse import urlencode
from urllib.request import urlopen
from zipfile import ZipFile

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


@cache
def get_schedule_for_teacher(filename: str, find: str) -> dict:
    """Функция для создания словаря путем парсинга excel таблицы"""

    book = load_workbook(
        filename=f"{os.getcwd()}/static/schedule/Pасписание/" + filename
    )
    ws = book.active
    # namep = (os.path.basename(filename))
    masstime = []
    masssur = []
    massgr = []
    masskb = []
    find = find
    for i in range(1, 18):
        for j in range(2, 34):
            lol = ws.cell(row=j, column=i).value
            if lol is None:
                lol = "@"
            if isinstance(lol, str):
                if find in lol:
                    if 3 <= j <= 8:
                        qaz = ws.cell(row=2, column=i).value
                    elif 10 <= j <= 15:
                        qaz = ws.cell(row=9, column=i).value
                    elif 17 <= j <= 22:
                        qaz = ws.cell(row=16, column=i).value
                    elif 24 <= j <= 29:
                        qaz = ws.cell(row=23, column=i).value
                    masstime.insert(0, ws.cell(row=j, column=2).value)
                    masssur.insert(0, lol.strip().replace("\n", " "))
                    massgr.insert(0, qaz)
                    masskb.insert(0, ws.cell(row=j, column=i + 1).value)

    lenght = len(masstime)

    pre_result = {}
    # заполняем словарь данными
    for i in range(lenght):
        pre_result[i] = {
            "Время: ": str(masstime[i]),
            "Предмет и преподаватель: ": str(masssur[i]),
            "Группа: ": str(massgr[i]),
            "Кабинет: ": str(masskb[i]),
        }

    # сортируем спарсенные данные по времени
    sotr_dict_key = sorted(
        pre_result, key=lambda x: int(pre_result[x]["Время: "].split(".")[0])
    )

    # перезаписываем отсортированные данные в словарь
    result = {}
    for i in range(lenght):
        result[i] = {
            "Время: ": str(masstime[sotr_dict_key[i]]),
            "Предмет и преподаватель: ": str(masssur[sotr_dict_key[i]]),
            "Группа: ": str(massgr[sotr_dict_key[i]]),
            "Кабинет: ": str(masskb[sotr_dict_key[i]]),
        }

    return result


@cache
def get_schedule_for_teacher_monday(filename: str, find: str) -> dict:
    """Функция для парсинга расписания для преподавателя в понедельник"""
    book = load_workbook(
        filename=f"{os.getcwd()}/static/schedule/Pасписание/" + filename
    )
    ws = book.active
    result = {}
    count = 0
    for i in range(1, 18):
        for j in range(2, 38):
            obj = str(ws.cell(row=j, column=i).value)
            if find in obj:
                group = "Не обнаружено"
                for x in range(1, 8):
                    if str(ws.cell(row=j - x, column=i).value)[:3].isdigit():
                        group = str(ws.cell(row=j - x, column=i).value)
                        break
                result[count] = {
                    "Время: ": str(ws.cell(row=j, column=2).value),
                    "Предмет и преподаватель: ": str(
                        ws.cell(row=j, column=i).value
                    ).replace("\n", " "),
                    "Группа: ": group,
                    "Кабинет: ": str(ws.cell(row=j, column=i + 1).value)
                    if "Россия" not in obj
                    else str(ws.cell(row=j - 1, column=i + 1).value),
                }
                count += 1

    sorted_schedule = dict(
        sorted(result.items(), key=lambda x: int(x[1]["Время: "].split(".")[0]))
    )
    return sorted_schedule


@cache
def get_schedule_for_group(filename: str, find: str) -> dict:
    """Функция для создания словаря путем парсинга excel таблицы"""
    book = load_workbook(
        filename=f"{os.getcwd()}/static/schedule/Pасписание/" + filename
    )
    ws = book.active
    SCHEDULE_LEN = 6
    result = {}
    for i in range(1, 18):
        for j in range(2, 34):
            obj = str(ws.cell(row=j, column=i).value)
            if obj == find:
                for iter in range(SCHEDULE_LEN):
                    result[iter] = {
                        "Время: ": str(ws.cell(row=j + iter + 1, column=2).value),
                        "Предмет и преподаватель: ": str(
                            ws.cell(row=j + iter + 1, column=i).value
                        ).replace("\n", " "),
                        "Группа: ": obj,
                        "Кабинет: ": str(ws.cell(row=j + iter + 1, column=i + 1).value),
                    }

    return result


@cache
def get_schedule_for_group_monday(filename: str, find: str) -> dict:
    """Функция для создания словаря путем парсинга excel таблицы"""
    book = load_workbook(
        filename=f"{os.getcwd()}/static/schedule/Pасписание/" + filename
    )
    ws = book.active
    SCHEDULE_LEN = 8
    result = {}
    for i in range(1, 18):
        for j in range(2, 34):
            obj = str(ws.cell(row=j, column=i).value)
            if obj == find:
                for iter in range(SCHEDULE_LEN):
                    result[iter] = {
                        "Время: ": str(ws.cell(row=j + iter + 1, column=2).value),
                        "Предмет и преподаватель: ": str(
                            ws.cell(row=j + iter + 1, column=i).value
                        ).replace("\n", " "),
                        "Группа: ": obj,
                        "Кабинет: ": str(ws.cell(row=j + iter + 1, column=i + 1).value)
                        if "Россия"
                        not in str(ws.cell(row=j + iter + 1, column=i).value).replace(
                            "\n", " "
                        )
                        else str(ws.cell(row=j + iter, column=i + 1).value),
                    }

    return result


def get_filename_list() -> list[str]:
    """Функция для получения доступных расписаний занятий"""

    list_name = os.listdir(os.getcwd() + "/static/schedule/Pасписание")

    result = []
    for i in list_name:
        result.append(i.replace(".xlsx", "").strip())

    return result


def check_on_update(url: str) -> bool:
    """Функция для проверки наличия обновлений"""

    st_accept = "text/html"

    st_useragent = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"

    headers = {"User-Agent": st_useragent, "Accept": st_accept}

    response = requests.get(url, headers=headers)

    src = response.text

    soup = BeautifulSoup(src, "html.parser")

    title_elements = soup.find_all(class_="listing-item__title")

    result_parse = []

    # Проходимся по найденным элементам и выводим содержимое каждого span
    for title_element in title_elements:
        span_element = title_element.find("span")
        if span_element:
            result_parse.append(span_element.text[:-5].rstrip())

    current_ver = get_filename_list()

    for i in result_parse:
        if i not in current_ver:
            return True

    return False


def del_from_folder() -> None:
    """Функция для удаления папки с расписаниями"""

    try:
        list = os.listdir(os.getcwd() + "/static/schedule/Pасписание")
        if list:
            shutil.rmtree(os.getcwd() + "/static/schedule/Pасписание")
    except:
        print("Folder not found")


def download_exel(url: str) -> None:
    """Функция для скачивания excel таблиц с расписанием"""

    base_url = "https://cloud-api.yandex.net/v1/disk/public/resources/download?"
    public_key = url

    # Получаем загрузочную ссылку
    final_url = base_url + urlencode(dict(public_key=public_key))
    response = requests.get(final_url)
    download_url = response.json()["href"]
    to_url = os.path.dirname(os.path.abspath(__file__))

    # записываем в переменную бинарные данные
    filedata = urlopen(download_url)
    datatowrite = filedata.read()

    # переводим бинарные данные в зип файл
    with open(to_url + "/static/schedule/qwe.zip", "wb") as f:
        f.write(datatowrite)

    # разархивируем архив
    with ZipFile(to_url + "/static/schedule/qwe.zip", "r") as zip_file:
        zip_file.extractall(to_url + "/static/schedule")
        os.remove(to_url + "/static/schedule/qwe.zip")

    file_list = os.walk(os.getcwd() + "/static/schedule/Pасписание")
    for root, dirs, files in file_list:
        for file in files:
            file_path = os.path.join(root, file)
            new_file_name = file[:-5].rstrip() + ".xlsx"
            new_file_path = os.path.join(root, new_file_name)
            os.rename(file_path, new_file_path)


class AsyncActionDownloadSchedule(threading.Thread):
    """Класс работающий не в основном потоке, который качает excel таблиц с рассписанием"""

    url_schedule = "https://disk.yandex.ru/d/VNdnX6hmveqJuw"

    def run(self):
        while True:
            # Проверяем наличие обновлений
            if not check_on_update(self.url_schedule):
                # таймаут в 15 минут
                time.sleep(900)
                continue

            del_from_folder()
            download_exel(url=self.url_schedule)

            # таймаут в 15 минут
            time.sleep(900)


# создаем экземпляр класса для скачивания excel таблиц
async_action_download_schedule = AsyncActionDownloadSchedule()
async_action_download_schedule.start()
